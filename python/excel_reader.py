"""
Python reading Excel files with OpenPyXL
Please Subscribe to Asim Code.
YouTube Channel: Asim Code
https://youtu.be/c2fg5qDHuFI
"""
import openpyxl
import openpyxl
book = openpyxl.load_workbook("SampleData.xlsx", data_only=True)
print(book.sheetnames)
for i in book.worksheets:
    print(i.title)

sheet = book["SalesOrders"]
print(sheet.max_row)
print(sheet.max_column)
print(sheet.cell(row=3, column=3).value)
